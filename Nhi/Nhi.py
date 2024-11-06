import os
from openpyxl import load_workbook

# Thư mục nguồn chứa các file Excel
source_folder = r'D:\aipdtest'
# Thư mục đích để lưu các file đã chỉnh sửa
destination_folder = r'D:\done'

# Tạo thư mục đích nếu chưa tồn tại
if not os.path.exists(destination_folder):
    os.makedirs(destination_folder)

# Lặp qua tất cả các file trong thư mục nguồn
for filename in os.listdir(source_folder):
    if filename.endswith('.xlsx'):
        # Đường dẫn tới file nguồn
        file_path = os.path.join(source_folder, filename)
        
        # Mở file Excel với openpyxl
        workbook = load_workbook(file_path)
        sheet = workbook.active

        # Tìm và thay thế "TỔNG GIÁ MUA" trong file
        found = False
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value == "TỔNG GIÁ MUA":
                    cell.value = "TỔNG GIÁ MUA (bao gồm VAT)"
                    # Chỉnh độ cao của hàng
                    sheet.row_dimensions[cell.row].height = 50  # Đặt độ cao mong muốn
                    found = True
                    break  # Thoát vòng lặp nội bộ khi tìm thấy
            if found:
                break  # Thoát vòng lặp chính sau khi tìm thấy

        # Đường dẫn lưu file vào thư mục đích
        new_file_path = os.path.join(destination_folder, filename)
        
        # Lưu file đã chỉnh sửa
        workbook.save(new_file_path)
        print(f"Đã lưu file: {new_file_path}")
