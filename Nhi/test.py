from openpyxl import load_workbook

# Đường dẫn tới file Excel
file_path = r'D:\aipdtest\TT00001.xlsx'

# Mở file Excel với openpyxl
workbook = load_workbook(file_path)

# Lấy sheet đầu tiên
sheet = workbook.active

# Duyệt qua các ô để tìm "TỔNG GIÁ MUA" và thay thế hàng đầu tiên tìm thấy
found = False
for row in sheet.iter_rows():
    for cell in row:
        if cell.value == "TỔNG GIÁ MUA":
            cell.value = "TỔNG GIÁ MUA (bao gồm VAT)"
            # Chỉnh độ cao của hàng
            sheet.row_dimensions[cell.row].height = 50  # Thay 25 bằng chiều cao mong muốn
            found = True
            break  # Thoát khỏi vòng lặp nội bộ sau khi tìm thấy
    if found:
        break  # Thoát khỏi vòng lặp chính sau khi tìm thấy

# Lưu file với tên mới để giữ nguyên file gốc
modified_file_path = r'D:\aipdtest\TT00001_modified_single_replace.xlsx'
workbook.save(modified_file_path)

print("File đã được lưu tại:", modified_file_path)
